from django.shortcuts import render, redirect, get_object_or_404
from api import models
from . import forms
from django.db.models import Count
from datetime import datetime, timedelta
import json
from django.utils import timezone
from django.db.models import Q
from django.http import Http404
from django.core.paginator import Paginator
from datetime import date
from django.db.models import Sum
from decimal import Decimal
import os
import cloudinary
import cloudinary.uploader
from openpyxl import Workbook
from django.conf import settings
from openpyxl.styles import PatternFill, Border, Side

# Vista para el index.html
def dashboard(request):
    
    total_usuarios = models.Usuario.objects.count()
    productos_activos = models.Producto.objects.filter(estado=True).count()
    total_ventas = models.Venta.objects.count()
    citas_pendientes = models.Cita.objects.filter(estado=False).count()
    
    
    ingresos_mensuales_citas = models.Cita.objects.filter(estado=True) \
        .values('fecha_cita__year', 'fecha_cita__month') \
        .annotate(ingreso_total=Sum('costo_cita')) \
        .order_by('fecha_cita__year', 'fecha_cita__month')
    meses_citas = [f"{ingreso['fecha_cita__year']}-{ingreso['fecha_cita__month']:02d}" for ingreso in ingresos_mensuales_citas]
    ingresos_citas = [float(ingreso['ingreso_total']) for ingreso in ingresos_mensuales_citas] 
    
    resultados_ventas = models.Venta.objects.values('fecha_venta__year', 'fecha_venta__month') \
        .annotate(ganancia_total=Sum('total')) \
        .order_by('fecha_venta__year', 'fecha_venta__month')
    meses_ventas = [f"{resultado['fecha_venta__year']}-{resultado['fecha_venta__month']:02d}" for resultado in resultados_ventas]
    ganancias_por_mes = [float(resultado['ganancia_total']) if isinstance(resultado['ganancia_total'], Decimal) else resultado['ganancia_total'] for resultado in resultados_ventas]
    
    meses_completos_2024 = [f"{año}-{mes:02d}" for año in range(2024, 2025) for mes in range(1, 13)]
    
    ingresos_completos_citas = [
        ingresos_citas[meses_citas.index(month)] if month in meses_citas else 0 for month in meses_completos_2024
    ]
    
    ganancias_completas_2024 = [
        ganancias_por_mes[meses_ventas.index(month)] if month in meses_ventas else 0 for month in meses_completos_2024
    ]
    
    ingresos_totales_por_mes = [ingresos_completos_citas[i] + ganancias_completas_2024[i] for i in range(len(meses_completos_2024))]

    datos_ingresos_totales = {
        'meses_completos_2024': meses_completos_2024,
        'ingresos_totales_por_mes': ingresos_totales_por_mes
    }

    datos_ingresos_totales_json = json.dumps(datos_ingresos_totales)
    
    
    return render(request, 'veterinaria/index.html', {
        'total_usuarios': total_usuarios,
        'productos_activos': productos_activos,
        'total_ventas': total_ventas,
        'citas_pendientes': citas_pendientes,
        'datos_ingresos_totales_json': datos_ingresos_totales_json
    })

def dashboard2(request):
    return render(request, 'veterinaria/index2.html')

def dashboard3(request):
    return render(request, 'veterinaria/index3.html')

def dashboard4(request):
    return render(request, 'veterinaria/borrar.html')


def widgets(request):
    
    fecha_actual = timezone.now()
    
    mes_actual = fecha_actual.month
    anio_actual = fecha_actual.year
    
    total_usuarios_mes = models.Usuario.objects.filter(
        fecha_registro__month=mes_actual,
        fecha_registro__year=anio_actual
    ).count()
    
    total_mascotas_mes = models.Mascota.objects.filter(
        fecha_inscripcion__month = mes_actual,
        fecha_inscripcion__year = anio_actual
    ).count()
    
    total_ventas_mes = models.Venta.objects.filter(
        fecha_venta__month = mes_actual,
        fecha_venta__year = anio_actual
    ).count()
    
    total_usuarios_completos = models.Usuario.objects.filter(
        Q(nombres__isnull=False) & ~Q(nombres='') &
        Q(apellidos__isnull=False) & ~Q(apellidos='') &
        Q(correo__isnull=False) & ~Q(correo='') &
        Q(username__isnull=False) & ~Q(username='') &
        Q(fecha_nacimiento__isnull=False) &
        Q(foto__isnull=False) & ~Q(foto='') &
        Q(telefono__isnull=False) & ~Q(telefono='') &
        Q(direccion__isnull=False) & ~Q(direccion='')
    ).count()

    total_usuarios = models.Usuario.objects.count()
    
    total_usuarios_con_mascota = models.Usuario.objects.filter(
        mascota__isnull=False  
    ).distinct().count()
    
    total_mascotas = models.Mascota.objects.count()
    
    total_servicios = models.Servicio.objects.count()
    
    total_veterinarios = models.Veterinario.objects.count()
    
    total_categorias_servicios = models.CategoriaServicio.objects.count()
    
    total_servicios_veterinarios = models.ServicioVeterinario.objects.count()
    
    total_categorias_productos = models.CategoriaProducto.objects.count()
    
    total_productos = models.Producto.objects.count()
    
    productos_unicos = models.DetalleVenta.objects.values('producto').distinct()

    total_productos_vendidos = productos_unicos.count()
    
    productos_activos = models.Producto.objects.filter(estado=True).count()

    productos_inactivos = models.Producto.objects.filter(estado=False).count()
    
    total_ventas = models.Venta.objects.count()
    
    total_metodos_pago = models.MetodoPago.objects.count()
    
    total_carritos = models.Carrito.objects.count()
    
    total_citas = models.Cita.objects.count()
    
    citas_pendientes = models.Cita.objects.filter(estado=False).count()

    citas_completadas = models.Cita.objects.filter(estado=True).count()
    
    return render(request, 'veterinaria/pages/widgets.html', {
        'total_usuarios': total_usuarios,
        'total_usuarios_mes': total_usuarios_mes,
        'total_usuarios_completos': total_usuarios_completos,
        'total_usuarios_con_mascota': total_usuarios_con_mascota,
        'total_servicios': total_servicios,
        'total_veterinarios': total_veterinarios,
        'total_categorias_servicios': total_categorias_servicios,
        'total_servicios_veterinarios': total_servicios_veterinarios,
        'total_mascotas': total_mascotas,
        'total_mascotas_mes': total_mascotas_mes,
        'total_categorias_productos': total_categorias_productos,
        'total_productos': total_productos,
        'total_ventas': total_ventas,
        'total_productos_vendidos': total_productos_vendidos,
        'productos_activos': productos_activos,
        'productos_inactivos': productos_inactivos,
        'total_ventas_mes': total_ventas_mes,
        'total_metodos_pago': total_metodos_pago,
        'total_carritos': total_carritos,
        'total_citas': total_citas,
        'citas_pendientes': citas_pendientes,
        'citas_completadas': citas_completadas,
    })


def chartjs(request):
    
    ##### ESPECIES DE MASCOTAS
    species = ['Perro', 'Gato', 'Otro']
    totals = [
        models.Mascota.objects.filter(especie = 'Perro').count(),
        models.Mascota.objects.filter(especie = 'Gato').count(),
        models.Mascota.objects.exclude(especie__in=['Perro', 'Gato']).count()
    ]
    species_json = json.dumps(species)
    totals_json = json.dumps(totals)
    
    
    ##### TOTAL DE ESPECIES DE MASCOTAS
    especies_all = models.Mascota.objects.values('especie').distinct()
    species_list = []
    totals_list = []
    for especie_all in especies_all:
        especie_name = especie_all['especie']
        total = models.Mascota.objects.filter(especie=especie_name).count()
        species_list.append(especie_name)
        totals_list.append(total)
    species_all_json = json.dumps(species_list)
    totals_all_json = json.dumps(totals_list)
    
    
    ##### EDADES DE LAS MASCOTAS
    edades = models.Mascota.objects.values('edad')
    edades_data = []
    for edad in edades:
        edades_data.append(edad['edad'])
    edades_data_json = json.dumps(edades_data)
    
    total_mascotas = models.Mascota.objects.count()
    total_mascotas_all = json.dumps(total_mascotas)


    ##### FECHAS DE NACIMIENTO DE MASCOTAS
    frecuencia = models.Mascota.objects.values('fecha_nacimiento__year').annotate(count=Count('mascota_id')).order_by('fecha_nacimiento__year')
    fechas = [str(item['fecha_nacimiento__year']) for item in frecuencia]
    frecuencias = [item['count'] for item in frecuencia]
    if frecuencia:
        min_year = frecuencia.first()['fecha_nacimiento__year']
        max_year = frecuencia.last()['fecha_nacimiento__year']
    else:
        min_year = max_year = datetime.now().year
    fechas_all = json.dumps(fechas)
    frecuencias_all = json.dumps(frecuencias)
    min_year_all = json.dumps(min_year)
    max_year_all = json.dumps(max_year)
    
    
    ##### GENERO DE LAS MASCOTAS
    hembras = models.Mascota.objects.filter(genero='Hembra').count()
    machos = models.Mascota.objects.filter(genero='Macho').count()
    datos_genero = [
        {"name": "Hembra", "value": hembras},
        {"name": "Macho", "value": machos}
    ]
    datos_genero_json = json.dumps(datos_genero)
    
    
    ##### FECHAS DE INSCRIPCION DE MASCOTAS
    mascotas_por_fecha = models.Mascota.objects.values('fecha_inscripcion__year', 'fecha_inscripcion__month').annotate(cantidad=Count('mascota_id')).order_by('fecha_inscripcion__year', 'fecha_inscripcion__month')
    fechas = [f"{mascota_id['fecha_inscripcion__year']}-{mascota_id['fecha_inscripcion__month']:02d}" for mascota_id in mascotas_por_fecha]
    cantidades = [mascota_id['cantidad'] for mascota_id in mascotas_por_fecha]
    all_months = [f"{year}-{month:02d}" for year in range(2024, 2025) for month in range(1, 13)]
    all_cantidades = [cantidades[fechas.index(month)] if month in fechas else 0 for month in all_months]
    data = {
        'fechas': all_months,
        'cantidades': all_cantidades
    }
    data_json = json.dumps(data)
    
    
    ##### EXCEL PARA REGISTRO DE MASCOTAS
    wb = Workbook()
    ws = wb.active
    ws.title = "Inscripciones de Mascotas"

    ws.append(["Fecha", "Cantidad"])

    for fecha, cantidad in zip(all_months, all_cantidades):
        ws.append([fecha, cantidad])

    tmp_dir = os.path.join(settings.BASE_DIR, 'tmp')  
    os.makedirs(tmp_dir, exist_ok=True)  

    excel_file = os.path.join(tmp_dir, 'mascotas_inscripcion.xlsx')
    
    wb.save(excel_file)

    upload_result = cloudinary.uploader.upload(excel_file, resource_type="raw")
    
    download_url = upload_result['secure_url']
    
    os.remove(excel_file)

    
    return render(request, 'veterinaria/pages/charts/chartjs.html', {
        'species_json': species_json,
        'totals_json': totals_json,
        'species_all_json': species_all_json,
        'totals_all_json': totals_all_json,
        'edades_data_json': edades_data_json,
        'total_mascotas_all': total_mascotas_all,
        'fechas_all': fechas_all,
        'frecuencias_all': frecuencias_all,
        'min_year_all': min_year_all,
        'max_year_all': max_year_all,
        'datos_genero_json': datos_genero_json,
        'data_json': data_json,
        'download_url': download_url
    })
    
    
def flot(request):
    
    ##### FECHAS DE REGISTRO DE USUARIOS
    usuarios_por_fecha = models.Usuario.objects.values('fecha_registro__year', 'fecha_registro__month').annotate(cantidad=Count('usuario')).order_by('fecha_registro__year', 'fecha_registro__month')
    fechas = [f"{usuario['fecha_registro__year']}-{usuario['fecha_registro__month']:02d}" for usuario in usuarios_por_fecha]
    cantidades = [usuario['cantidad'] for usuario in usuarios_por_fecha]
    all_months = [f"{year}-{month:02d}" for year in range(2024, 2025) for month in range(1, 13)]
    all_cantidades = [cantidades[fechas.index(month)] if month in fechas else 0 for month in all_months]
    data = {
        'fechas': all_months,
        'cantidades': all_cantidades
    }
    data_json = json.dumps(data)
    
    
    ###### EXCEL PARA EL REGISTRO DE USUARIOS
    wb = Workbook()
    ws = wb.active
    ws.title = "Inscripciones de Mascotas"

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fecha = "Fecha"
    header_cantidad = "Cantidad"
    
    ws.append([header_fecha, header_cantidad])
    
    for col_num in range(1, 3):  
        cell = ws.cell(row=1, column=col_num)
        cell.fill = yellow_fill  
        cell.border = border  
        cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')

    for row_num, (fecha, cantidad) in enumerate(zip(all_months, all_cantidades), start=2):
        ws.cell(row=row_num, column=1, value=fecha).border = border 
        ws.cell(row=row_num, column=2, value=cantidad).border = border  

    ws.column_dimensions['A'].width = 15  
    ws.column_dimensions['B'].width = 10  

    tmp_dir = os.path.join(settings.BASE_DIR, 'tmp')
    os.makedirs(tmp_dir, exist_ok=True)

    excel_file = os.path.join(tmp_dir, 'mascotas_inscripcion.xlsx')

    wb.save(excel_file)

    upload_result = cloudinary.uploader.upload(excel_file, resource_type="raw")
    
    download_url = upload_result['secure_url']
    
    os.remove(excel_file)
    
    
    
    
    ##### EDADES DE LOS USUARIOS
    usuarios = models.Usuario.objects.filter(fecha_nacimiento__isnull=False)
    edades = []
    for usuario in usuarios:
        edad = date.today().year - usuario.fecha_nacimiento.year
        if date.today().month < usuario.fecha_nacimiento.month or (date.today().month == usuario.fecha_nacimiento.month and date.today().day < usuario.fecha_nacimiento.day):
            edad -= 1  
        edades.append(edad)
    intervalos = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
    conteo_edades = {f"{intervalo}-{intervalo+9}": 0 for intervalo in intervalos if intervalo < 100}
    for edad in edades:
        for intervalo in intervalos:
            if intervalo <= edad < intervalo + 10:
                rango = f"{intervalo}-{intervalo + 9}"
                conteo_edades[rango] += 1
                break
    intervalos_json = json.dumps(list(conteo_edades.keys()))
    conteo_json = json.dumps(list(conteo_edades.values()))
    
    ##### SERVICIOS DE CITAS
    servicios = models.Cita.objects.values('servicio_id__nombre').annotate(cantidad=Count('servicio_id')).order_by('-cantidad')
    nombres_servicios = [servicio['servicio_id__nombre'] for servicio in servicios]
    cantidades_servicios = [servicio['cantidad'] for servicio in servicios]
    nombres_json = json.dumps(nombres_servicios)
    cantidades_servicios__json = json.dumps(cantidades_servicios)
    
    
    ##### CITAS PENDIENTES DE LA SEMANA
    hoy = date.today()
    lunes = hoy - timedelta(days=hoy.weekday())  
    domingo = lunes - timedelta(days=7) 
    citas = models.Cita.objects.filter(fecha_cita__gte=domingo, fecha_cita__lte=hoy, estado=False)
    citas_por_dia = {lunes + timedelta(days=i): 0 for i in range(7)}  
    for cita in citas:
        if cita.fecha_cita in citas_por_dia:
            citas_por_dia[cita.fecha_cita] += 1
    fechas_semana = [cita.strftime('%d/%m/%Y') for cita in citas_por_dia.keys()]
    cantidad_citas = list(citas_por_dia.values())
    fechas_semana_json = json.dumps(fechas_semana)
    cantidad_json = json.dumps(cantidad_citas)
    
    
    ##### HORARIOS DE LAS CITAS
    citas_por_horario = models.Cita.objects.values('horario_id__hora') \
        .annotate(citas_count=Count('horario_id')) \
        .order_by('-citas_count') 
    horarios = [cita['horario_id__hora'].strftime("%H:%M") for cita in citas_por_horario]
    cantidad_citas_horarios = [cita['citas_count'] for cita in citas_por_horario]
    data_horarios = {
        'horarios': horarios,
        'cantidad_citas_horarios': cantidad_citas_horarios
    }
    data_horarios_json = json.dumps(data_horarios)
    
    
    ##### ESTADO DE LAS CITAS
    estados_citas = ['Pendiente', 'Completada']
    cantidades_estados = [
        models.Cita.objects.filter(estado=False).count(), 
        models.Cita.objects.filter(estado=True).count()    
    ]
    estados_citas_json = json.dumps(estados_citas)
    cantidades_estados_json = json.dumps(cantidades_estados)
    
    
    ##### INGRESOS DE LAS CITAS
    ingresos_mensuales_citas = models.Cita.objects.filter(estado=True) \
        .values('fecha_cita__year', 'fecha_cita__month') \
        .annotate(ingreso_total=Sum('costo_cita')) \
        .order_by('fecha_cita__year', 'fecha_cita__month')
    meses_citas = [f"{ingreso['fecha_cita__year']}-{ingreso['fecha_cita__month']:02d}" for ingreso in ingresos_mensuales_citas]
    ingresos_citas = [float(ingreso['ingreso_total']) for ingreso in ingresos_mensuales_citas] 
    meses_completos_citas = [f"{year}-{month:02d}" for year in range(2024, 2025) for month in range(1, 13)]
    ingresos_completos_citas = [ingresos_citas[meses_citas.index(month)] if month in meses_citas else 0 for month in meses_completos_citas]
    data_ingresos_citas = {
        'meses_completos_citas': meses_completos_citas,
        'ingresos_completos_citas': ingresos_completos_citas
    }
    data_ingresos_citas_json = json.dumps(data_ingresos_citas)
    
    ##### EXCEL PARA INGRESOS DE LAS CITAS
    wb_citas = Workbook()
    ws_citas = wb_citas.active
    ws_citas.title = "Ingresos de Citas"

    amarillo_fondo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
    borde_citas = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    encabezado_mes = "Mes"
    encabezado_ingreso_total = "Ingreso Total"
    
    ws_citas.append([encabezado_mes, encabezado_ingreso_total])

    for col_num in range(1, 3):  
        cell = ws_citas.cell(row=1, column=col_num)
        cell.fill = amarillo_fondo  
        cell.border = borde_citas  
        cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')

    for row_num, (mes, ingreso) in enumerate(zip(meses_completos_citas, ingresos_completos_citas), start=2):
        ws_citas.cell(row=row_num, column=1, value=mes).border = borde_citas 
        ws_citas.cell(row=row_num, column=2, value=ingreso).border = borde_citas 

    ws_citas.column_dimensions['A'].width = 15 
    ws_citas.column_dimensions['B'].width = 15  

    tmp_dir_citas = os.path.join(settings.BASE_DIR, 'tmp')
    os.makedirs(tmp_dir_citas, exist_ok=True)

    archivo_excel_citas = os.path.join(tmp_dir_citas, 'ingresos_citas.xlsx')

    wb_citas.save(archivo_excel_citas)

    resultado_subida_citas = cloudinary.uploader.upload(archivo_excel_citas, resource_type="raw")
    
    url_descarga_citas = resultado_subida_citas['secure_url']
    
    os.remove(archivo_excel_citas)

    
    return render(request, 'veterinaria/pages/charts/flot.html', {
        'data_json': data_json,
        'intervalos': intervalos_json,
        'conteo': conteo_json,
        'fechas_semana': fechas_semana_json,
        'cantidad_citas': cantidad_json,
        'estados_citas': estados_citas_json,
        'cantidades_estados': cantidades_estados_json,
        'data_ingresos_citas_json': data_ingresos_citas_json,
        'data_horarios_json': data_horarios_json,
        'nombres_servicios': nombres_json,
        'cantidades_servicios': cantidades_servicios__json,
        'download_url': download_url,
        'url_descarga_citas': url_descarga_citas
    })
    

def inline(request):
    
    ##### LISTA DE PRODUCTOS INACTIVOS
    productos_inactivos = models.Producto.objects.filter(estado=False)
    
    
    ##### LISTA DE PRODUCTOS ACTIVOS
    productos_activos = models.Producto.objects.filter(estado=True)
    
    
    ##### STOCK ALTO
    productos_stock_alto = models.Producto.objects.filter(stock__gte=10, estado=True).values('nombre', 'stock')
    productos_stock_alto = productos_stock_alto.order_by('-stock')
    productos_nombres = [producto['nombre'] for producto in productos_stock_alto]
    productos_stock = [producto['stock'] for producto in productos_stock_alto]
    data_stock_alto = {
        'productos_nombres': productos_nombres,
        'productos_stock': productos_stock
    }
    data_stock_json = json.dumps(data_stock_alto)
    
    
    ##### STOCK BAJO
    productos_stock_bajo = models.Producto.objects.filter(stock__lt=4, estado=True).values('nombre', 'stock')
    productos_stock_bajo = productos_stock_bajo.order_by('stock')
    productos_nombres_bajo = [producto['nombre'] for producto in productos_stock_bajo]
    productos_stock_bajo = [producto['stock'] for producto in productos_stock_bajo]
    data_stock_bajo = {
        'productos_nombres_bajo': productos_nombres_bajo,
        'productos_stock_bajo': productos_stock_bajo
    }
    data_stock_bajo_json = json.dumps(data_stock_bajo)
    
    
    ##### TOTAL DE VENTAS
    ventas_por_fecha = models.Venta.objects.values('fecha_venta__year', 'fecha_venta__month') \
        .annotate(cantidad=Count('id')) \
        .order_by('fecha_venta__year', 'fecha_venta__month')
    fechas_ventas_productos = [f"{venta['fecha_venta__year']}-{venta['fecha_venta__month']:02d}" for venta in ventas_por_fecha]
    cantidades_ventas = [venta['cantidad'] for venta in ventas_por_fecha]
    all_monthsH = [f"{year}-{month:02d}" for year in range(2024, 2025) for month in range(1, 13)]
    all_cantidadesH = [cantidades_ventas[fechas_ventas_productos.index(month)] if month in fechas_ventas_productos else 0 for month in all_monthsH]
    data_ventas_productos = {
        'all_monthsH': all_monthsH,
        'all_cantidadesH': all_cantidadesH
    }
    data_ventas_productos_json = json.dumps(data_ventas_productos)
    
    ##### EXCEL PARA VENTAS
    wb_ventas = Workbook()
    ws_ventas = wb_ventas.active
    ws_ventas.title = "Ventas de Productos"

    amarillo_fondo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    encabezado_mes = "Mes"
    encabezado_cantidad_ventas = "Cantidad de Ventas"
    
    ws_ventas.append([encabezado_mes, encabezado_cantidad_ventas])

    for col_num in range(1, 3):  
        cell = ws_ventas.cell(row=1, column=col_num)
        cell.fill = amarillo_fondo  
        cell.border = borde  
        cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')

    for row_num, (mes, cantidad) in enumerate(zip(all_monthsH, all_cantidadesH), start=2):
        ws_ventas.cell(row=row_num, column=1, value=mes).border = borde 
        ws_ventas.cell(row=row_num, column=2, value=cantidad).border = borde 

    ws_ventas.column_dimensions['A'].width = 15
    ws_ventas.column_dimensions['B'].width = 20

    tmp_dir_ventas = os.path.join(settings.BASE_DIR, 'tmp')
    os.makedirs(tmp_dir_ventas, exist_ok=True)

    archivo_excel_ventas = os.path.join(tmp_dir_ventas, 'ventas_productos.xlsx')

    wb_ventas.save(archivo_excel_ventas)

    resultado_subida_ventas = cloudinary.uploader.upload(archivo_excel_ventas, resource_type="raw")

    url_descarga_ventas = resultado_subida_ventas['secure_url']

    os.remove(archivo_excel_ventas)
    
    
    ##### GANANCIAS DE PRODUCTOS POR MES
    resultados_ventas = models.Venta.objects.values('fecha_venta__year', 'fecha_venta__month') \
        .annotate(ganancia_total=Sum('total')) \
        .order_by('fecha_venta__year', 'fecha_venta__month')
    meses_ventas = [f"{resultado['fecha_venta__year']}-{resultado['fecha_venta__month']:02d}" for resultado in resultados_ventas]
    ganancias_por_mes = [float(resultado['ganancia_total']) if isinstance(resultado['ganancia_total'], Decimal) else resultado['ganancia_total'] for resultado in resultados_ventas]
    meses_completos_2024 = [f"{año}-{mes:02d}" for año in range(2024, 2025) for mes in range(1, 13)]
    ganancias_completas_2024 = [ganancias_por_mes[meses_ventas.index(mes)] if mes in meses_ventas else 0 for mes in meses_completos_2024]
    datos_ventas_mensuales = {
        'meses_completos_2024': meses_completos_2024,
        'ganancias_completas_2024': ganancias_completas_2024
    }
    datos_ventas_mensuales_json = json.dumps(datos_ventas_mensuales)
    
    ###### EXCEL PARA LAS GANANCIAS
    

    
    return render(request, 'veterinaria/pages/charts/inline.html', {
        'productos_inactivos': productos_inactivos,
        'productos_activos': productos_activos,
        'data_stock_json': data_stock_json,
        'data_stock_bajo_json': data_stock_bajo_json,
        'data_ventas_productos_json': data_ventas_productos_json,
        'datos_ventas_mensuales_json': datos_ventas_mensuales_json,
        'url_descarga_ventas': url_descarga_ventas
    })
    

def uplot(request):
    return render(request, 'veterinaria/pages/charts/uplot.html')


def tableUsuarios(request):
    nombre_usuario = request.GET.get('nombre_usuario', '')
    if nombre_usuario:
        usuarios = models.Usuario.objects.filter(nombres__icontains = nombre_usuario)
    else:
        usuarios = models.Usuario.objects.all()
    paginator = Paginator(usuarios, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/usuarios.html', {'page_obj': page_obj})

def verUsuario(request, usuario):
    usuario = get_object_or_404(models.Usuario, usuario=usuario)
    return render(request, 'veterinaria/pages/tables/ver_usuario.html', {'usuario': usuario})

def crearUsuario(request):
    return render(request, 'veterinaria/pages/tables/crear_usuario.html')

def editarUsuario(request, usuario_id):
    return render(request, 'veterinaria/pages/tables/editar_usuario.html')

def eliminarUsuario(request, usuario_id):
    return render(request, 'veterinaria/pages/tables/eliminar_usuario.html')


def tableEspecialidad(request):
    nombre_servicio_veterinario = request.GET.get('nombre_servicio_veterinario', '')
    if nombre_servicio_veterinario:
        especialidades = models.ServicioVeterinario.objects.filter(nombre__icontains = nombre_servicio_veterinario)
    else:
        especialidades = models.ServicioVeterinario.objects.all()
    paginator = Paginator(especialidades, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/especialidad.html', {'page_obj': page_obj})

def verEspecialidad(request, Servicio_veterinario_id):
    especialidad = get_object_or_404(models.ServicioVeterinario, Servicio_veterinario_id=Servicio_veterinario_id)
    return render(request, 'veterinaria/pages/tables/ver_especialidad.html', {'especialidad': especialidad})

def crearEspecialidad(request):
    if request.method == 'POST':
        form = forms.EspecialidadVeterinarioForm(request.POST)
        if form.is_valid():
            form.save() 
            return redirect('table-especialidad')  
    else:
        form = forms.EspecialidadVeterinarioForm()

    return render(request, 'veterinaria/pages/tables/crear_especialidad.html', {'form': form})

def editarEspecialidad(request, Servicio_veterinario_id):
    especialidad = get_object_or_404(models.ServicioVeterinario, Servicio_veterinario_id=Servicio_veterinario_id)
    if request.method == 'POST':
        form = forms.EspecialidadVeterinarioForm(request.POST, instance=especialidad)
        if form.is_valid():
            form.save()  
            return redirect('table-especialidad')  
    else:
        form = forms.EspecialidadVeterinarioForm(instance=especialidad) 

    return render(request, 'veterinaria/pages/tables/editar_especialidad.html', {'form': form})

def eliminarEspecialidad(request, Servicio_veterinario_id):
    especialidad = get_object_or_404(models.ServicioVeterinario, Servicio_veterinario_id=Servicio_veterinario_id)
    
    if request.method == 'POST':
        especialidad.delete()  
        return redirect('table-usuarios')  
    
    return render(request, 'veterinaria/pages/tables/eliminar_especialidad.html', {'especialidad': especialidad})


def tableVeterinarios(request):
    nombre_veterinario = request.GET.get('nombre_veterinario', '')
    if nombre_veterinario:
        veterinarios = models.Veterinario.objects.filter(nombres__icontains = nombre_veterinario)
    else:
        veterinarios = models.Veterinario.objects.all()
    paginator = Paginator(veterinarios, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/veterinarios.html', {'page_obj': page_obj})

def verVeterinario(request, veterinario_id):
    veterinario = get_object_or_404(models.Veterinario, veterinario_id=veterinario_id)
    return render(request, 'veterinaria/pages/tables/ver_veterinarios.html', {'veterinario': veterinario})

def crearVeterinario(request):
    if request.method == 'POST':
        form = forms.VeterinarioForm(request.POST)
        if form.is_valid():
            form.save() 
            return redirect('table-veterinarios')  
    else:
        form = forms.VeterinarioForm()

    return render(request, 'veterinaria/pages/tables/crear_veterinario.html', {'form': form})

def editarVeterinario(request, veterinario_id):
    veterinario = get_object_or_404(models.Veterinario, veterinario_id=veterinario_id)
    if request.method == 'POST':
        form = forms.VeterinarioForm(request.POST, instance=veterinario)
        if form.is_valid():
            form.save()  
            return redirect('table-veterinarios')  
    else:
        form = forms.VeterinarioForm(instance=veterinario) 

    return render(request, 'veterinaria/pages/tables/editar_veterinario.html', {'form': form})

def eliminarVeterinario(request, veterinario_id):
    veterinario = get_object_or_404(models.Veterinario, veterinario_id=veterinario_id)
    
    if request.method == 'POST':
        veterinario.delete()  
        return redirect('table-veterinarios')  
    
    return render(request, 'veterinaria/pages/tables/eliminar_veterinario.html', {'veterinario': veterinario})


def tableCategoriasServicios(request):
    nombre_categoria_servicio = request.GET.get('nombre_categoria_servicio', '')
    if nombre_categoria_servicio:
        categoriasServicios = models.CategoriaServicio.objects.filter(nombre__icontains = nombre_categoria_servicio)
    else:
        categoriasServicios = models.CategoriaServicio.objects.all()
    paginator = Paginator(categoriasServicios, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/categorias_servicios.html', {'page_obj': page_obj})

def verCategoriaServicios(request, categoria_servicio_id):
    categoria_servicio = get_object_or_404(models.CategoriaServicio, categoria_servicio_id=categoria_servicio_id)
    return render(request, 'veterinaria/pages/tables/ver_categorias_servicios.html', {'categoria_servicio': categoria_servicio})

def crearCategoriaServicios(request):
    if request.method == 'POST':
        form = forms.CategoriaServicioForm(request.POST)
        if form.is_valid():
            form.save() 
            return redirect('table-categorias-servicios')  
    else:
        form = forms.CategoriaServicioForm()

    return render(request, 'veterinaria/pages/tables/crear_categoria_servicio.html', {'form': form})

def editarCategoriaServicios(request, categoria_servicio_id):
    categoria_servicio = get_object_or_404(models.CategoriaServicio, categoria_servicio_id=categoria_servicio_id)
    if request.method == 'POST':
        form = forms.CategoriaServicioForm(request.POST, instance=categoria_servicio)
        if form.is_valid():
            form.save()  
            return redirect('table-veterinarios')  
    else:
        form = forms.CategoriaServicioForm(instance=categoria_servicio) 

    return render(request, 'veterinaria/pages/tables/editar_categoria_servicios.html', {'form': form})


def tableServicios(request):
    nombre_servicio = request.GET.get('nombre_servicio', '')
    if nombre_servicio:
        servicios = models.Servicio.objects.filter(nombre__icontains = nombre_servicio)
    else:
        servicios = models.Servicio.objects.all()
    paginator = Paginator(servicios, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/servicios.html', {'page_obj': page_obj})

def verServicio(request, servicio_id):
    servicio = get_object_or_404(models.Servicio, servicio_id=servicio_id)
    return render(request, 'veterinaria/pages/tables/ver_servicio.html', {'servicio': servicio})

def crearServicio(request):
    if request.method == 'POST':
        form = forms.ServicioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('table-servicios')  
    else:
        form = forms.ServicioForm()

    return render(request, 'veterinaria/pages/tables/crear_servicio.html', {'form': form})

def editarServicio(request, servicio_id):
    servicio = get_object_or_404(models.Servicio, servicio_id=servicio_id)
    if request.method == 'POST':
        form = forms.ServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            return redirect('table-servicios') 
    else:
        form = forms.ServicioForm(instance=servicio)

    return render(request, 'veterinaria/pages/tables/editar_servicio.html', {'form': form})


def tableDetallesVeterinaria(request):
    detallesVeterinaria = models.DetalleVeterinaria.objects.all()
    paginator = Paginator(detallesVeterinaria, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/detalles_veterinaria.html', {'page_obj': page_obj})

def verDetalleVeterinaria(request, detalle_veterinaria_id):
    detalle_veterinaria = get_object_or_404(models.DetalleVeterinaria, detalle_veterinaria_id=detalle_veterinaria_id)
    return render(request, 'veterinaria/pages/tables/ver_detalle_veterinaria.html', {'detalle_veterinaria': detalle_veterinaria})

def crearDetalleVeterinaria(request):
    if request.method == 'POST':
        form = forms.DetalleVeterinariaForm(request.POST)
        if form.is_valid():
            form.save() 
            return redirect('table-detalles-veterinaria')  
    else:
        form = forms.DetalleVeterinariaForm()

    return render(request, 'veterinaria/pages/tables/crear_detalle_veterinaria.html', {'form': form})

def editarDetalleVeterinaria(request, detalle_veterinaria_id):
    detalle_veterinaria = get_object_or_404(models.DetalleVeterinaria, detalle_veterinaria_id=detalle_veterinaria_id)
    if request.method == 'POST':
        form = forms.DetalleVeterinariaForm(request.POST, instance=detalle_veterinaria)
        if form.is_valid():
            form.save()  
            return redirect('table-detalles-veterinaria')  
    else:
        form = forms.DetalleVeterinariaForm(instance=detalle_veterinaria) 

    return render(request, 'veterinaria/pages/tables/editar_detalle_veterinaria.html', {'form': form})


def tableCategoriasProductos(request):
    nombre_categoria_producto = request.GET.get('nombre_categoria_producto', '')
    if nombre_categoria_producto:
        categoriasProductos = models.CategoriaProducto.objects.filter(nombre__icontains = nombre_categoria_producto)
    else:
        categoriasProductos = models.CategoriaProducto.objects.all()
    paginator = Paginator(categoriasProductos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/categorias_productos.html', {'page_obj': page_obj})

def verCategoriaProducto(request, categoria_producto_id):
    categoria_producto = get_object_or_404(models.CategoriaProducto, categoria_producto_id=categoria_producto_id)
    return render(request, 'veterinaria/pages/tables/ver_categoria_producto.html', {'categoria_producto': categoria_producto})

def crearCategoriaProducto(request):
    if request.method == 'POST':
        form = forms.CategoriaProductoForm(request.POST)
        if form.is_valid():
            form.save() 
            return redirect('table-categorias-productos') 
    else:
        form = forms.CategoriaProductoForm()

    return render(request, 'veterinaria/pages/tables/crear_categoria_producto.html', {'form': form})

def editarCategoriaProducto(request, categoria_producto_id):
    categoria_producto = get_object_or_404(models.CategoriaProducto, categoria_producto_id=categoria_producto_id)
    if request.method == 'POST':
        form = forms.CategoriaProductoForm(request.POST, instance=categoria_producto)
        if form.is_valid():
            form.save()  
            return redirect('table-categorias-productos')  
    else:
        form = forms.CategoriaProductoForm(instance=categoria_producto)

    return render(request, 'veterinaria/pages/tables/editar_categoria_producto.html', {'form': form})


def tableProductos(request):
    nombre_producto = request.GET.get('nombre_producto', '')
    if nombre_producto:
        productos = models.Producto.objects.filter(nombre__icontains = nombre_producto)
    else:
        productos = models.Producto.objects.all()
    paginator = Paginator(productos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/productos.html', {'page_obj': page_obj})

def verProducto(request, producto_id):
    producto = get_object_or_404(models.Producto, producto_id=producto_id)
    return render(request, 'veterinaria/pages/tables/ver_producto.html', {'producto': producto})

def crearProducto(request):
    if request.method == 'POST':
        form = forms.ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save() 
            return redirect('table-productos')  
    else:
        form = forms.ProductoForm()

    return render(request, 'veterinaria/pages/tables/crear_producto.html', {'form': form})

def editarProducto(request, producto_id):
    producto = get_object_or_404(models.Producto, producto_id=producto_id)
    if request.method == 'POST':
        form = forms.ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()  
            return redirect('table-productos') 
    else:
        form = forms.ProductoForm(instance=producto)

    return render(request, 'veterinaria/pages/tables/editar_producto.html', {'form': form})


def tableMetodosPago(request):
    nombre_metodo_pago = request.GET.get('nombre_metodo_pago', '') 
    if nombre_metodo_pago:
        metodosPago = models.MetodoPago.objects.filter(nombre__icontains=nombre_metodo_pago)
    else:
        metodosPago = models.MetodoPago.objects.all()
    paginator = Paginator(metodosPago, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/metodos_pago.html', {'page_obj': page_obj})

def verMetodoPago(request, metodo_id):
    metodo_pago = get_object_or_404(models.MetodoPago, metodo_id=metodo_id)
    return render(request, 'veterinaria/pages/tables/ver_metodo_pago.html', {'metodo_pago': metodo_pago})

def crearMetodoPago(request):
    if request.method == 'POST':
        form = forms.MetodoPagoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('table-metodos-pago')  
    else:
        form = forms.MetodoPagoForm()

    return render(request, 'veterinaria/pages/tables/crear_metodo_pago.html', {'form': form})

def editarMetodoPago(request, metodo_id):
    metodo_pago = get_object_or_404(models.MetodoPago, metodo_id=metodo_id)
    if request.method == 'POST':
        form = forms.MetodoPagoForm(request.POST, instance=metodo_pago)
        if form.is_valid():
            form.save()  
            return redirect('table-metodos-pago') 
    else:
        form = forms.MetodoPagoForm(instance=metodo_pago)

    return render(request, 'veterinaria/pages/tables/editar_metodo_pago.html', {'form': form})


def tableCarritos(request):
    nombre_carrito = request.GET.get('nombre_carrito', '')
    if nombre_carrito:
        carritos = models.Carrito.objects.filter(usuario__nombres__icontains = nombre_carrito)
    else:
        carritos = models.Carrito.objects.all()
    paginator = Paginator(carritos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/carritos.html', {'page_obj': page_obj})

def verCarrito(request, carrito_id):
    carrito = get_object_or_404(models.Carrito, id=carrito_id) 
    return render(request, 'veterinaria/pages/tables/ver_carrito.html', {'carrito': carrito})


def tableProductosCarrito(request):
    nombre_producto_carrito = request.GET.get('nombre_producto_carrito', '')
    if nombre_producto_carrito:
        productos_carrito = models.ProductoCarrito.objects.filter(
            Q(producto__nombre__icontains=nombre_producto_carrito) | 
            Q(carrito__usuario__nombres__icontains=nombre_producto_carrito)
        )
    else:
        productos_carrito = models.ProductoCarrito.objects.all()
    paginator = Paginator(productos_carrito, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/productos_carrito.html', {'page_obj': page_obj})

def verProductosCarrito(request, carrito_id):
    carrito = get_object_or_404(models.Carrito, id=carrito_id)  
    productos_carrito = models.ProductoCarrito.objects.filter(carrito=carrito)  
    if not productos_carrito:
        raise Http404("No hay productos en este carrito.")
    return render(request, 'veterinaria/pages/tables/ver_productos_carrito.html', {'carrito': carrito, 'productos_carrito': productos_carrito})


def tableVentas(request):
    nombre_venta = request.GET.get('nombre_venta', '')
    if nombre_venta:
        ventas = models.Venta.objects.filter(usuario__nombres__icontains = nombre_venta)
    else:
        ventas = models.Venta.objects.all()
    paginator = Paginator(ventas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/ventas.html', {'page_obj': page_obj})

def verVenta(request, venta_id):
    venta = get_object_or_404(models.Venta, id=venta_id)  
    return render(request, 'veterinaria/pages/tables/ver_venta.html', {'venta': venta})


def tableDetallesVenta(request):
    nombre_detalle_venta = request.GET.get('nombre_detalle_venta', '')
    if nombre_detalle_venta:
        detalles_venta = models.DetalleVenta.objects.filter(
            Q(venta__usuario__nombres__icontains = nombre_detalle_venta) |
            Q(producto__nombre__icontains = nombre_detalle_venta)
        )
    else:
        detalles_venta = models.DetalleVenta.objects.all()
    paginator = Paginator(detalles_venta, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/detalles_venta.html', {'page_obj': page_obj})

def verDetalleVenta(request, venta_id):
    venta = get_object_or_404(models.Venta, id=venta_id)  
    detalles_venta = models.DetalleVenta.objects.filter(venta=venta)  
    if not detalles_venta:
        raise Http404("No hay detalles para esta venta.")
    return render(request, 'veterinaria/pages/tables/ver_detalle_venta.html', {'venta': venta, 'detalles_venta': detalles_venta})


def tableMascotas(request):
    nombre_mascota = request.GET.get('nombre_mascota', '')
    if nombre_mascota:
        mascotas = models.Mascota.objects.filter(nombre__icontains = nombre_mascota)
    else:
        mascotas = models.Mascota.objects.all()
    paginator = Paginator(mascotas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/mascotas.html', {'page_obj': page_obj})

def verMascota(request, mascota_id):
    mascota = get_object_or_404(models.Mascota, mascota_id=mascota_id)
    return render(request, 'veterinaria/pages/tables/ver_mascota.html', {'mascota': mascota})

def crearMascota(request):
    if request.method == 'POST':
        form = forms.MascotaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('table-mascotas')  
    else:
        form = forms.MascotaForm()

    return render(request, 'veterinaria/pages/tables/crear_mascota.html', {'form': form})

def editarMascota(request, mascota_id):
    mascota = get_object_or_404(models.Mascota, mascota_id=mascota_id)
    if request.method == 'POST':
        form = forms.MascotaForm(request.POST, request.FILES, instance=mascota)
        if form.is_valid():
            form.save()  
            return redirect('table-mascotas')  
    else:
        form = forms.MascotaForm(instance=mascota)

    return render(request, 'veterinaria/pages/tables/editar_mascota.html', {'form': form})


def tableHorarios(request):
    horarios = models.Horario.objects.all()
    paginator = Paginator(horarios, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/horarios.html', {'page_obj': page_obj})

def verHorario(request, horario_id):
    horario = get_object_or_404(models.Horario, horario_id=horario_id)
    return render(request, 'veterinaria/pages/tables/ver_horario.html', {'horario': horario})

def crearHorario(request):
    if request.method == 'POST':
        form = forms.HorarioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('table-horarios')  
    else:
        form = forms.HorarioForm()

    return render(request, 'veterinaria/pages/tables/crear_horario.html', {'form': form})

def editarHorario(request, horario_id):
    horario = get_object_or_404(models.Horario, horario_id=horario_id)
    if request.method == 'POST':
        form = forms.HorarioForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()  
            return redirect('table-horarios')  
    else:
        form = forms.HorarioForm(instance=horario)

    return render(request, 'veterinaria/pages/tables/editar_horario.html', {'form': form})


def tableCitas(request):
    nombre_cita = request.GET.get('nombre_cita', '')
    if nombre_cita:
        citas = models.Cita.objects.filter(usuario_id__nombres__icontains = nombre_cita)
    else:
        citas = models.Cita.objects.all()
    paginator = Paginator(citas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/citas.html', {'page_obj': page_obj})

def verCita(request, cita_id):
    cita = get_object_or_404(models.Cita, cita_id=cita_id)
    return render(request, 'veterinaria/pages/tables/ver_cita.html', {'cita': cita})

def crearCita(request):
    if request.method == 'POST':
        form = forms.CitaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('table-citas')
    else:
        form = forms.CitaForm()

    return render(request, 'veterinaria/pages/tables/crear_cita.html', {'form': form})

def editarCita(request, cita_id):
    cita = get_object_or_404(models.Cita, cita_id=cita_id)
    if request.method == 'POST':
        form = forms.CitaForm(request.POST, instance=cita)
        if form.is_valid():
            form.save()
            return redirect('table-citas')
    else:
        form = forms.CitaForm(instance=cita)

    return render(request, 'veterinaria/pages/tables/editar_cita.html', {'form': form})


def tableHistorialMascotas(request):
    nombre_historial_mascota = request.GET.get('nombre_historial_mascota', '')
    if nombre_historial_mascota:
        historiales = models.HistorialMascota.objects.filter(mascota_id__nombre__icontains = nombre_historial_mascota)
    else:
        historiales = models.HistorialMascota.objects.all()
    paginator = Paginator(historiales, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'veterinaria/pages/tables/historial_mascotas.html', {'page_obj': page_obj})

def verHistorialMascota(request, historial_id):
    historial = get_object_or_404(models.HistorialMascota, historial_id=historial_id) 
    return render(request, 'veterinaria/pages/tables/ver_historial_mascota.html', {'historial': historial})